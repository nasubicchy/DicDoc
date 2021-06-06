"""EXCELSheetを任意のPathに作成"""
import openpyxl
import os
import tkinter as tk
import tkinter.ttk
import tkinter.filedialog as fldg
import tkinter.messagebox
import sys
from functools import partial # 関数を部分適用するためのモジュール
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side # セルをデザインするためのモジュール
import copy # 深いコピーのためのモジュール

"""=========関数定義=================="""
# #==================================
# # システム関係
# #==================================
def app_exit(): # app終了
    ans = tkinter.messagebox.askyesno(
        '終了確認',
        '自動保存はされません。\n保存したい場合は「保存」か「App終了」を押してください。\n本当に辞書登録を終了してもいいですか？'
    )
    if ans == False:
        pass
    else:
        print("アプリケーションを終了します。")
        sys.exit()

# #==================================
# # xlsxファイル新規作成・既存ファイル編集
# #==================================
def create_xlsx():
    typ = [('EXCEL',"*.xlsx")]
    global FileName1 #Global変数へ編集可能に
    FileName1 = fldg.asksaveasfilename(filetypes = typ)
    if FileName1 != "":
        if FileName1[-5:] != ".xlsx":
            FileName1 = FileName1 + ".xlsx"
    # print(f'保存先は{FileName1}')
    global Edt_Mode
    Edt_Mode = 0 # 編集モード指定
    GUI_inpath.destroy() #　コマンド終了後ウインドウを閉じるように設定
    return FileName1, Edt_Mode

def edit_xlsx():
    typ = [('EXCEL',"*.xlsx")]
    global FileName1 #Global変数へ編集可能に
    FileName1 = fldg.askopenfilename(filetypes = typ) # 編集するファイルの名前を表示
    global Edt_Mode
    Edt_Mode = 1 # 編集モード指定

    # Excelシートの内容読み込み
    r_wb = openpyxl.load_workbook(FileName1)
    r_sh = r_wb.active
    global Dic_Exist
    # Dic_Exist = list() # 初期値：読み込み辞書
    column = 1 # 行指定の初期値
    for row in range(1,r_sh.max_row + 1): # 1行目からSheetの最大行まで読み込む
        # 各列の要素を行用オブジェクトに記入していく
        dic_elem = list() # 初期値＆初期化：行要素用辞書オブジェクト
        wrd1 = str(r_sh.cell(row,column).value)
        if wrd1 is None: # 白紙ファイルを編集用として開いてしまったときの対応(無いと”None”がCellに追加されてしまう)
            break
        wrd2 = str(r_sh.cell(row,column+1).value)
        wrd3 = str(r_sh.cell(row,column+2).value)
        wrd4 = str(r_sh.cell(row,column+3).value)
        wrd5 = str(r_sh.cell(row,column+4).value)
        wrd6 = str(r_sh.cell(row,column+5).value)
        dic_elem.append(wrd1)
        dic_elem.append(wrd2)
        dic_elem.append(wrd3)
        dic_elem.append(wrd4)
        dic_elem.append(wrd5)
        dic_elem.append(wrd6)
        # Noneでセルに表示されるところをブランクに変更する
        elem_num = 0
        for wrd in dic_elem:
            if wrd == 'None':
                dic_elem[elem_num] = ''
                elem_num += 1
            else:
                elem_num += 1
        Dic_Exist.append(dic_elem)
    # 1行目のタイトル行を削除(dicオブジェクトに誤って形状されないため)
    Dic_Exist.pop(0)
    # Dic_Existのなかにある目次行の削除
    dic_row_num_error = list()
    dic_row_num = 0
    for row in Dic_Exist:
        if row[1] == '':
            dic_row_num_error.append(dic_row_num)
            dic_row_num += 1
        else:
            dic_row_num += 1
    dlt_count = 0 # 辞書から要素を削除した回数→indexがずれることへの調整
    for dlt_index in dic_row_num_error:
            Dic_Exist.pop(dlt_index - dlt_count)
            dlt_count += 1

    GUI_inpath.destroy() #　コマンド終了後ウインドウを閉じるように設定
    return FileName1, Edt_Mode, Dic_Exist

# #==================================
# # 文字列読み込み・削除
# #==================================
def read_str(dic,mode,event=None):
    wrd1 = str(txt_box1.get()) # 単語
    wrd2 = str(txt_box2.get()) # よみ
    wrd3 = str(txt_box3.get('1.0','end-1c')) # 意味・説明
    wrd4 = str(txt_box4.get()) # ジャンル
    if wrd4 == KIND[0]: # ジャンルが初期値の場合は空欄を返す
        wrd4 = ''
    wrd5 = str(txt_box5.get()) # 備考
    wrd6 = str(txt_box6.get()) # 参考URL
    txt_box1.delete(0,tk.END)
    txt_box2.delete(0,tk.END)
    txt_box3.delete('1.0',tk.END)
    txt_box4.delete(0,tk.END)
    txt_box5.delete(0,tk.END)
    txt_box6.delete(0,tk.END)
    # print(f'単語：{wrd1}\n単語の説明：{wrd3}') #テスト用
    # 単語をリストへ格納していく
    global Dic_Input
    global Read_Once
    if mode == 1: # 元のEXCELから読み込む
        if Read_Once == 0:
            Dic_Input = copy.deepcopy(dic) # 読み込んだ辞書リストオブジェクトを深いコピー
            Read_Once = 1
        else:
            pass
    elif mode == 0:
        pass
    dict_elem = list()
    dict_elem.append(wrd1)
    dict_elem.append(wrd2)
    dict_elem.append(wrd3)
    dict_elem.append(wrd4)
    dict_elem.append(wrd5)
    dict_elem.append(wrd6)
    Dic_Input.append(dict_elem)
    # print(f'単語：{wrd1}\n単語のよみ：{wrd2}\n単語の説明：{wrd3}') #テスト用
    print(f'関数read_str使用時のDic_Inputの長さは{len(Dic_Input)}') #テスト用
    return Dic_Input, Read_Once

def delete_str(event=None):
    txt_box1.delete(0,tk.END)
    txt_box2.delete(0,tk.END)
    txt_box3.delete('1.0',tk.END)
    txt_box4.delete(0,tk.END)
    txt_box5.delete(0,tk.END)
    txt_box6.delete(0,tk.END)

# #==================================
# # Sorting関数
# #==================================
def sorting(order,dic): # sortingの順番と、sortingの対象    
    # 2次元Listを縦から読んで、List内をindex順に読み取る
    # ひらがなの順番を参考に、各要素の末尾2列に順番IDを付与
    dic_row_num = 0 # 辞書オブジェクトの単語番号
    dic_row_num_error = list() # 「よみ」にエラーがあった辞書の列番号を格納した行列
    dic_sort1 = copy.deepcopy(dic) # 引数で読み込んだリストを深いコピー
    for row in dic_sort1:
        # sort_orderの"単語"の1文字目を順番と一致していれば、そのindex番号を返す
        sort_row_num = 0 # sort_orderの行番号の初期値
        detect_wrd = False # ひらがな順番から単語発見できたか否かの判断の初期値
        # 単語にSorting番号を付与していない場合、付与を開始
        if detect_wrd == False:
            for sort_row in order:
                if detect_wrd == False:
                    sort_col_num = 0 # sort_orderの列番号の初期値
                    for sort_col in sort_row:
                        if row[1].startswith(sort_col) == True:
                            dic_sort1[dic_row_num].append(sort_row_num)
                            dic_sort1[dic_row_num].append(sort_col_num)
                            detect_wrd = True
                            break
                        sort_col_num += 1
                else:
                    break
                sort_row_num += 1
        # sort_orderをすべて読みこんだ上で、エラーがあるよみがあれば、辞書オブジェクトの単語番号を追加する。
            if detect_wrd == False:
                dic_row_num_error.append(dic_row_num)
                detect_wrd = True
            else:
                pass
        dic_row_num += 1
    # print(f'辞書の要素数は{len(dic)}') #テスト用
    # print(f'エラーのある読みを含む単語は{len(dic_row_num_error)}')#テスト用
    # print(f'エラーのある読みを含む単語のindexは\n{dic_row_num_error}')#テスト用

    # 順番リストに無い文字は、エラーウインドウを表示し、辞書から削除
    if len(dic_row_num_error) != 0:
        error_win = tk.Tk()
        error_win.withdraw()
        cfm = tkinter.messagebox.showinfo(
            'エラー',
            '無効な入力がありました。\n\n「よみ」に直接入力の英数字・かな入力以外を記入した単語は辞書から削除しました。'
        )
        if cfm == 'ok':
            dlt_count = 0 # 辞書から要素を削除した回数→indexがずれることへの調整
            for dlt_index in dic_row_num_error:
                print(f'削除した登録単語の単語番号は{dlt_index}') # テスト用
                dic_sort1.pop(dlt_index - dlt_count)
                dlt_count += 1
            error_win.destroy()
    else:
        pass
    # 上記で付与した順番IDを基にSorting
    dic_sort2 = copy.deepcopy(dic_sort1) # index番号を付加したリストを深いコピー
    for i in range(0,len(dic_sort2)-1): #len関数は個数で返すので要素数読み込みのために−1しておく、※1行目はタイトル行なので、始点をずらして1から開始
        for j in range(i+1,len(dic_sort2)):
            if dic_sort2[i][-2] > dic_sort2[j][-2]: # 最後から二番目の要素（ひらがな行を参考にSorting）
                tmp = dic_sort2[j]
                dic_sort2[j] = dic_sort2[i]
                dic_sort2[i] = tmp
            elif dic_sort2[i][-2] == dic_sort2[j][-2]:
                if dic_sort2[i][-1] > dic_sort2[j][-1]: # 最後から1番目の要素（ひらがな行内のaioeoを参考にSorting）
                    tmp = dic_sort2[j]
                    dic_sort2[j] = dic_sort2[i]
                    dic_sort2[i] = tmp

    return dic_sort2

# #==================================
# # Sorting後にジャンル・その他を記入する関数
# #==================================
def design(order,dic,head): # sortingの順番と、sortingの対象、タイトル行の設定
    dic_design1 = copy.deepcopy(dic) # 読み込んだリストを深いコピー
    
    ''' 目次行を追加する '''
    # title_append = False # タイトル行の追加がされたかどうか。※各50音で追加されたらリセットされる
    
    for sort_row in order:
        for sort_col in sort_row:
            title_row = [str(sort_col),'','','','','','s1','s2'] # 追加するためのタイトル行を生成
            dic_row_num = 0 # 辞書オブジェクトの単語番号
            for dic_row in dic_design1:
                if dic_row[1].startswith(sort_col) == True:
                    # 該当行の直前にタイトル行を挿入
                    dic_design1.insert(dic_row_num,title_row)
                    break
                dic_row_num += 1
        

    ''' タイトル行がdicオブジェクトの1行目になければ、追加する '''
    for i in range(0,len(dic_design1[0])):
        # print(dic[0][i]) #テスト用
        # print(head[i]) #テスト用
        if dic_design1[0][i] != head[i]:
            if i == len(dic_design1[0])-1:
                dic_design1.insert(0,head) # タイトル行を先頭に追加
            else:
                continue
        else:
            break

    return dic_design1




# #==================================
# # EXCELオブジェクト作成・記入
# #==================================
def output_excel(dic,wb):
    # Excelbookオブジェクト作成
    global Wb_Obj
    if wb is None: # 既にEXCELオブジェクトが作成済かどうか判断
        Wb_Obj = openpyxl.Workbook()
    else:
        Wb_Obj = copy.deepcopy(wb)
    sh_obj = Wb_Obj.active
    row_num = 0 # 読み込んだ辞書オブジェクトの列番号（単語番号）
    # 2次元Listを縦から読んで、List内をindex順に読み取って、Sheetに書き込んでいく
    for row in dic:
        col_num = 0
        for CellData in row:
            if col_num < len(row)-2:
                sh_obj.cell(row_num+1,col_num+1).value = CellData
                col_num += 1
            else:
                col_num += 1
        row_num += 1

    '''Sheetの全体デザイン変更'''
    # タイトル行（1行目）の固定
    sh_obj.freeze_panes = 'A2'
    # シート全体の罫線を消す
    sh_obj.sheet_view.showGridLines = False
    # すべてのセルを選択していく
    row_num = 0 # 読み込んだ辞書オブジェクトの行番号（単語番号）
    for row in sh_obj:
        col_num = 0 # 読み込んだ辞書オブジェクトの列番号
        for cell in row:
            if col_num == 5:
                sh_obj[cell.coordinate].alignment = Alignment(
                    shrinkToFit= True, #「縮小して全体表示」を可能に
                    vertical='top',
                    )
                cell.hyperlink = cell.value
            else:
                sh_obj[cell.coordinate].alignment = Alignment(
                    wrapText=True, #「折り返して表示」を可能に
                    vertical='top'
                )
            cell.border = BORDER1
            # タイトル行だけ文字の形変更
            if row_num == 0:
                cell.font = EX_FONT1
                cell.fill = PatternFill(
                    patternType='solid',
                    fgColor=TITLE_CELL_COLER1
                )
            col_num += 1
        # タイトル行のみデザイン変更
        if row[1].value == '' and row[2].value == '' and row[3].value == '' and row[4].value == '' and row[5].value == '':
            row[0].font = EX_FONT2
            for i in range(0,6):
                row[i].fill = PatternFill(
                    patternType='solid',
                    fgColor=TITLE_CELL_COLER2
                )
            for i in range(0,6):
                if i == 0:
                    row[i].border = BORDER2
                elif i == 5:
                    row[i].border = BORDER4
                else:
                    row[i].border = BORDER3
        row_num += 1

    # 列幅設定
    for col_index in COL_WIDTHS.keys():
        sh_obj.column_dimensions[col_index].width = COL_WIDTHS[col_index]
    
    return Wb_Obj

# #==================================
# # 保存用ボタン
# #==================================
def save(order,dic,dic_ex,mode,filename,wb,head,event=None):
    # print(order)# テスト用
    # print(dic)# テスト用
    # print(dic_ex)# テスト用
    if Edt_Mode == 1:
        if Read_Once == 1:
            dic_bsort = copy.deepcopy(dic)
        elif Read_Once == 0:
            dic_bsort = copy.deepcopy(dic_ex)
    else:
        dic_bsort = copy.deepcopy(dic)
    dic_sorted = sorting(order,dic_bsort)
    # print(dic_sorted) # テスト用
    dic_designed = design(order,dic_sorted,head)
    output_excel(dic_designed,wb)
    # print('単語登録して、順番を治した辞書オブジェクトは') # テスト用
    # print(dic_designed) # テスト用

    # Excelオブジェクト保存
    if mode == 0 or 1:
        Wb_Obj.save(filename)
    elif mode is None:
        pass

# #==================================
# # App終了
# #==================================
def ext_win():

    GUI_input.destroy()
    return

"""=========ウインドウ用設定パラメーター=================="""  
FONT1 = ("system",13)
FONT2 = ('system',18)
EX_FONT1 = Font(
    name = '游ゴシック',
    size = 18,
    bold = True
    )
EX_FONT2 = Font(
    name = '游ゴシック',
    size = 12,
    bold = True
    )
SIDE1 = Side(
    style='thin',
    color='000000'
)
SIDE2 = Side(
    style=None,
    color='000000'
)
SIDE3 = Side(
    style='thick',
    color='000000'
)
BORDER1 = Border(
    left= SIDE1,
    right= SIDE1,
    top= SIDE1,
    bottom= SIDE1
)
BORDER2 = Border(
    left= SIDE1,
    right= SIDE2,
    top= SIDE1,
    bottom= SIDE1
)
BORDER3 = Border(
    left= SIDE2,
    right= SIDE2,
    top= SIDE1,
    bottom= SIDE1
)
BORDER4 = Border(
    left= SIDE2,
    right= SIDE1,
    top= SIDE1,
    bottom= SIDE1
)
TITLE_CELL_COLER1 = "AA8866"
TITLE_CELL_COLER2 = "ceedff"

HEAD = [
    '単語',
    'よみ',
    '意味',
    'ジャンル',
    '備考メモ',
    '参考URL',
    't1',
    't2'
]
COL_WIDTHS = {
    "A":10,
    "B":10,
    "C":80,
    "D":40,
    "E":40,
    "F":40
}


"""=========初期値====================================="""  
FileName1 = None # Globalオブジェクトを空で作成
Edt_Mode = None
Wb_Obj = None
Read_Once = 0 # 重要。読み込んだファイルを編集ファイルに設定済みかどうかを判断。
# 辞書用List
Dic_Input = list()
Dic_Exist = list()

"""=========sortingの順番====================================="""  
sort_order = (
    ('A','a','B','b','C','c','D','d','E','e','F','f',\
        'G','g','H','h','I','i','J','j','K','k','L','l','M','m',\
            'N','n','O','o','P','p','Q','q','R','r','S','s','T','t',\
                'U','u','V','v','W','w','X','x','Y','y','Z','z'),
    ('あ','い','う','ゔ','え','お'),
    ('か','が','き','ぎ','く','ぐ','け','げ','こ','ご'),
    ('さ','ざ','し','じ','す','ず','せ','ぜ','そ','ぞ'),
    ('た','だ','ち','ぢ','つ','づ','て','で','と','ど'),
    ('な','に','ぬ','ね','の'),
    ('は','ば','ぱ','ひ','び','ぴ','ふ','ぶ','ぷ','へ','べ','ぺ','ほ','ぼ','ぽ'),
    ('ま','み','む','め','も'),
    ('や','ゆ','よ'),
    ('ら','り','る','れ','ろ'),
    ('わ','を','ん'),
    ('1','2','3','4','5','6','7','8','9','0')
    )
"""=========単語カテゴライズのList=====================================""" 
KIND = [
    'ジャンルを選択可能です。記入も可。',
    'CLW07',
    'CLW08',
    'CLS-50',
    'CLS-50H',
    'CLS-50Ⅱ',
    'ZT',
    'HID3',
    'HID4',
    'HID5',
    'HID6',
    'VHL',
    'STK',
    'CONV',
    '制御関係',
    'MTL',
    'MTM',
    'MTH',
    '施工関係',
    '営業関係'
]

"""=================================="""
"""=========処理開始=================="""
"""=================================="""  
if __name__ == "__main__":
    #==================================
    # ファイル保存先指定用のGUI
    #==================================
    # ウインドウ構成
    GUI_inpath = tk.Tk()
    GUI_inpath.title('Create a new entry')
    GUI_inpath.resizable(False,False)
    GUI_inpath.geometry('250x100')

    # 説明ウインドウ
    msg1 = tk.Message(
        text = 'Select edit mode.',
        width = 180,
        fg = 'black',
        font = FONT2,
        # bg = 'gray'
        )
    # 新規作成ボタン
    btn_new = tk.Button(
        text = 'Create a new',
        font = FONT1,
        command = create_xlsx,
        # bg = 'white',
        fg = 'black'
        )
    # 既存ファイル編集ボタン
    btn_edt = tk.Button(
        text = 'Edit exist file',
        font = FONT1,
        command = edit_xlsx,
        # bg = 'white',
        fg = 'black'
        )

    msg1.pack()
    btn_new.pack()
    btn_edt.pack()


    # ファイル作成ウインドウ実行
    GUI_inpath.protocol('WM_DELETE_WINDOW',app_exit) # xボタンを押した際にプログラム終了
    GUI_inpath.mainloop()

    # print(f'読み込んだファイルは\n{Dic_Exist}\n') # テスト用
    print(f'Dic_Inputの長さは\n{len(Dic_Input)}\n') # テスト用
    print(f'Dic_Existの長さは\n{len(Dic_Exist)}\n') # テスト用


    #==================================
    # 辞書入力画面用のGUI
    #==================================
    # ウインドウ構成
    GUI_input = tk.Tk()
    GUI_input.title('Input word')
    GUI_input.geometry('450x340')
    GUI_input.resizable(
        width = False,
        height = False
    )

    # 単語入力Box
    lbl1 = tk.Label(
        text = '登録単語',
        fg = 'black'
        )
    txt_box1 = tk.Entry(
        width = 40,
        bg = 'white',
        fg = 'black'
        )
    # 単語よみ入力Box
    lbl2 = tk.Label(
        text = 'よみ',
        fg = 'black'
        )
    txt_box2 = tk.Entry(
        width = 40,
        bg = 'white',
        fg = 'black'
        )
    # 単語説明入力Box
    lbl3 = tk.Label(
    text = '単語の説明↓',
    fg = 'black'
    )
    txt_box3 = tk.Text(
        bg = 'white',
        fg = 'black'
    )
    # 単語ジャンル入力Box
    lbl4 = tk.Label(
        text = 'ジャンル',
        fg = 'black'
        )
    txt_box4 = tkinter.ttk.Combobox(
        state = 'normal',
        values = KIND,
        width = 30
        )
    txt_box4.set(KIND[0])
    # 単語の備考入力Box
    lbl5 = tk.Label(
        text = '備考メモ',
        fg = 'black'
        )
    txt_box5 = tk.Entry(
        width = 40,
        bg = 'white',
        fg = 'black'
        )
    # 参考URL入力Box
    lbl6 = tk.Label(
        text = '参考URL',
        fg = 'black'
        )
    txt_box6 = tk.Entry(
        width = 40,
        bg = 'white',
        fg = 'black'
        )

    # 登録ボタン
    btn_wrd1 = tk.Button(
        text = '登録',
        command = partial(read_str,Dic_Exist,Edt_Mode),
        fg = 'black'
        )
    # 削除ボタン
    btn_wrd2 = tk.Button(
        text = '削除',
        command = delete_str,
        fg = 'black'
        )
    # 終了ボタン
    btn_ext = tk.Button(
        text = '保存してApp終了',
        command = ext_win,
        fg = 'black'
        )

    # ウインドウコンテンツ配置
    lbl1.place(
        x = 10,
        y = 10
    )
    txt_box1.place(
        x = 70,
        y = 10
        ) 
    lbl2.place(
        x = 10,
        y = 40
    )
    txt_box2.place(
        x = 70,
        y = 40
        ) 
    lbl3.place(
        x = 10,
        y = 70
    )
    txt_box3.place(
        x = 10,
        y = 100,
        width = 430,
        height = 100,
        )
    lbl4.place(
        x = 10,
        y = 210
    )
    txt_box4.place(
        x = 70,
        y = 210
        ) 
    lbl5.place(
        x = 10,
        y = 240
    )
    txt_box5.place(
        x = 70,
        y = 240
        ) 
    lbl6.place(
        x = 10,
        y = 270
    )
    txt_box6.place(
        x = 70,
        y = 270
        ) 
    btn_wrd1.place(
        x = 60,
        y = 300
        )
    btn_wrd2.place(
        x = 130,
        y = 300
        )
    btn_ext.place(
        x = 300,
        y = 300
        )

    # メニューバー
    mbar_input = tk.Menu()
    mcom_input = tk.Menu(
        mbar_input,
        tearoff = 0
    )
    mcom_input.add_command(
        label = '登録',
        command = partial(read_str,Dic_Exist,Edt_Mode),
        accelerator = 'Ctrl-Q'
    )
    mcom_input.add_command(
        label = '削除',
        command = delete_str,
        accelerator = 'Ctrl-D'
    )
    mbar_input.add_cascade(
        label = '編集',
        menu = mcom_input
    )
    GUI_input['menu'] = mbar_input

    # ウインドウ上のキー入力での反応を設定（ショートカット）
    GUI_input.bind('<Control-q>',partial(read_str,Dic_Exist,Edt_Mode))
    GUI_input.bind('<Control-d>',delete_str)

    # 入力画面ウインドウ実行
    GUI_input.protocol('WM_DELETE_WINDOW',app_exit) # xボタンを押した際にプログラム終了
    GUI_input.mainloop()

    print(f'Dic_Inputの長さは\n{len(Dic_Input)}\n') # テスト用
    print(f'Dic_Existの長さは\n{len(Dic_Exist)}\n') # テスト用


 

    '''=================================='''
    ''' 辞書データを保存する作業'''
    '''=================================='''
    save(sort_order,Dic_Input,Dic_Exist,Edt_Mode,FileName1,Wb_Obj,HEAD)
 







